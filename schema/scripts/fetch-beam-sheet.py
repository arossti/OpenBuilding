#!/usr/bin/env python3
"""Pull every BEAM worksheet tab via the xlsx export endpoint.

Why not gviz? The gviz/tq?tqx=out:csv endpoint silently truncates large
tabs at the last contiguous data block (Footings & Slabs: 749 rows in the
sheet -> 317 lines from gviz, a 57% data loss that slipped past us in
session 3). The xlsx export returns the full workbook.

Dep: openpyxl. Run `pip3 install openpyxl` if not already installed.

Usage:
  python3 schema/scripts/fetch-beam-sheet.py             fetch default tab set
  python3 schema/scripts/fetch-beam-sheet.py Tab1 Tab2   fetch specific tabs
"""

import csv
import sys
import urllib.request
from pathlib import Path

import openpyxl  # type: ignore

SHEET_ID = "1LjOpDTjfGQvvfRGCpDb8KkHcUtHzUC5UbvfV-wXy13g"

DEFAULT_TABS = [
    "PROJECT",
    "Footings & Slabs",
    "Foundation Walls",
    "Structural Elements",
    "Ext. Walls",
    "Party Walls",
    "Cladding",
    "Windows",
    "Int. Walls",
    "Floors",
    "Ceilings",
    "Roof",
    "Garage",
    "REVIEW",
    "RESULTS",
    "Materials DB",
    "Categories",
    "Materials",
    "Glossary",
    "PROJECT_FIELDS",
    "Settings",
    "Version",
]

REPO_ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = REPO_ROOT / "docs" / "csv files from BEAM"
CACHE_XLSX = Path("/tmp") / "beam.xlsx"


def download_xlsx() -> Path:
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
    print(f"  downloading workbook -> {CACHE_XLSX}")
    with urllib.request.urlopen(url) as res:  # follows 307 redirect automatically
        data = res.read()
    if not data[:2] == b"PK":
        raise RuntimeError("response is not a zip/xlsx; sheet may not be publicly viewable")
    CACHE_XLSX.write_bytes(data)
    print(f"  wrote {len(data) / 1024:.0f} KB")
    return CACHE_XLSX


def export_tab(wb, tab: str, out_dir: Path) -> int:
    if tab not in wb.sheetnames:
        raise KeyError(f"tab not found in workbook")
    ws = wb[tab]
    out = out_dir / f"{tab}.csv"
    rows = 0
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            w.writerow(["" if v is None else v for v in row])
            rows += 1
    return rows


def main(argv):
    tabs = argv[1:] if len(argv) > 1 else DEFAULT_TABS
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = download_xlsx()
    # data_only=True reads cached formula values; read_only=True for memory efficiency
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ok = 0
    fail = 0
    for tab in tabs:
        try:
            rows = export_tab(wb, tab, OUT_DIR)
            print(f"  OK {tab.ljust(22)} {rows} rows")
            ok += 1
        except Exception as e:
            print(f"  FAIL {tab.ljust(22)} {e}")
            fail += 1
    print(f"\n{ok} ok, {fail} failed -> {OUT_DIR}")
    if fail:
        sys.exit(1)


if __name__ == "__main__":
    main(sys.argv)
